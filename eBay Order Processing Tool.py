import requests
import openpyxl
from datetime import datetime, timezone, timedelta
import tkinter as tk
from tkinter import filedialog, ttk
import json
import os
import platformdirs

#Definiere eine Klasse, die bei leerem Eingabefeld einen grauen Platzhaltertext anzeigt. Wenn das Eingabefeld den Fokus erh lt, verschwindet der Platzhaltertext. Wenn das Eingabefeld den Fokus verliert und leer ist, erscheint der Platzhaltertext wieder.
class PlaceholderEntry(ttk.Entry):
    """Eingabefeld mit grauem Platzhaltertext, der bei Fokus verschwindet."""
    
    def __init__(self, container, placeholder, *args, **kwargs):
        super().__init__(container, *args, **kwargs, style='Placeholder.TEntry')
        self.placeholder = placeholder
        self.insert('0', self.placeholder)
        self.bind('<FocusIn>', self._clear_placeholder)
        self.bind('<FocusOut>', self._add_placeholder)
        self.configure(foreground='grey')

    def _clear_placeholder(self, e):
        """Entfernt den Platzhalter bei Fokus."""
        if self.get() == self.placeholder:
            self.delete('0', 'end')
            self.configure(foreground='white')

    def _add_placeholder(self, e):
        """Fügt den Platzhalter wieder ein, wenn das Feld leer ist."""
        if not self.get():
            self.insert('0', self.placeholder)
            self.configure(foreground='grey')
    
    def get_value(self):
        """Gibt den aktuellen Wert zurück (leerer String falls nur Platzhalter)."""
        return '' if self.get() == self.placeholder else self.get()

# Konfigurationsdateipfad
def get_config_file_path():
    """
    Erstellt einen plattformunabh ngigen Konfigurationsdateipfad
    Verwendet den Skriptnamen als Konfigurationsdateiname, um eine einfache Identifizierung zu erm glichen
    """
    # Skriptname ohne Dateiendung extrahieren
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    app_author = "MeinUnternehmen"
    
    # Plattformspezifisches Anwendungsdatenverzeichnis abrufen
    data_dir = platformdirs.user_data_dir("EbayTools", app_author, ensure_exists=True)
    
    # Konfigurationsdateinamen erstellen
    config_file = f"{script_name}.json"
    config_path = os.path.join(data_dir, config_file)
    
    # Konfigurationspfad für Debugging ausgeben
    print(f"Konfigurationsdateipfad: {config_path}")
    return config_path

def load_config():
    """Lädt die gespeicherte Konfiguration"""
    config_file = get_config_file_path()
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Fehler beim Lesen der Konfigurationsdatei: {e}")
    return {
        "token": "",
        "days": "",
        "orders_limit": "",
        "excel_path": "",
        "worksheet_name": ""
    }   # Leeres Wörterbuch zurückgeben, wenn die Konfigurationsdatei nicht existiert oder ungültig ist

def save_config(config):
    """Speichert die aktuelle Konfiguration"""
    try:
        config_file = get_config_file_path()
        # Sicherstellen, dass das Verzeichnis existiert
        os.makedirs(os.path.dirname(config_file), exist_ok=True)
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"Fehler beim Speichern der Konfiguration: {e}")

"""
Hauptlogik zur Datenverarbeitung
1. Alle Bestellungen mit getOrders aus Fulfillment abrufen und dann mit getOrder jede Bestellung durchlaufen, um Informationen in einer Liste zu speichern
2. Die erhaltene Informationsliste weiterverarbeiten, stornierte Bestellungen entfernen
3. Entsprechend der Modellkennung verarbeiten
4. Die verarbeiteten Daten mit den Daten in der Excel-Datei vergleichen, Duplikate entfernen und schließlich in Excel schreiben
"""

# Eine Hauptfunktion erstellen, um die gesamte UI-Interaktion zu verarbeiten:
def main():
    # Konfiguration laden
    config = load_config()

    # Hauptfenster erstellen
    root = tk.Tk()
    root.title("eBay Bestellverarbeitungsprogramm")

    """
    Dateiauswahldialog für Excel-Datei
    
    Diese Funktion öffnet einen Dateiauswahldialog, mit dem der Benutzer
    eine Excel-Datei auswählen kann. Die ausgewählte Datei wird dann
    in das entsprechende Eingabefeld eingetragen.
    """
    def browse_file():
        filename = filedialog.askopenfilename()
        excel_entry.delete(0, tk.END)
        excel_entry.insert(0, filename)

    # Fenstergröße definieren
    width = 800
    height = 900
    root.geometry(
        f"{width}x{height}+{(root.winfo_screenwidth() - width) // 2}+{(root.winfo_screenheight() - height) // 2}")

    # Einen Rahmen zum Platzieren von UI-Elementen erstellen
    frame = tk.Frame(root)
    frame.pack(padx=20, pady=20)

    # Gewichtung der zweiten Spalte einstellen, damit sie sich horizontal ausdehnt
    frame.columnconfigure(1, weight=1)

    # Label für "eBay Zugriffstoken" erstellen
    token_label = tk.Label(frame, text="eBay access token:")
    token_label.grid(row=0, column=0, sticky="ew", padx=10, pady=0)

    # Eingabefeld für eBay-Zugriffstoken erstellen
    token_entry = tk.Text(frame, height=10)
    token_entry.grid(row=1, column=0, columnspan=2, sticky="ew", padx=10, pady=0)

    # Stil für Placeholder-Text erstellen
    style = ttk.Style()
    style.configure('Placeholder.TEntry', foreground='grey')
    
    # Label und Eingabefeld für die Anzahl der Bestellungstage erstellen
    days_frame = tk.Frame(frame)
    days_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=10, pady=10)
    days_frame.columnconfigure(1, weight=1)
    
    days_label = tk.Label(days_frame, text="Order days:")
    days_label.grid(row=0, column=0, sticky="w")
    
    days_entry = PlaceholderEntry(days_frame, "z.B. 3")
    if "days" in config and config["days"]:
        days_entry.delete(0, 'end')
        days_entry.insert(0, config["days"])
        days_entry.configure(foreground='white')
    days_entry.grid(row=0, column=1, sticky="ew", padx=(10, 0))

    # Label und Eingabefeld für die maximale Anzahl der Bestellungen erstellen
    orders_frame = tk.Frame(frame)
    orders_frame.grid(row=3, column=0, columnspan=2, sticky="ew", padx=10, pady=10)
    orders_frame.columnconfigure(1, weight=1)
    
    orders_label = tk.Label(orders_frame, text="Orders limit:")
    orders_label.grid(row=0, column=0, sticky="w")
    
    orders_entry = PlaceholderEntry(orders_frame, "z.B. 100")
    if "orders_limit" in config and config["orders_limit"]:
        orders_entry.delete(0, 'end')
        orders_entry.insert(0, config["orders_limit"])
        orders_entry.configure(foreground='white')
    orders_entry.grid(row=0, column=1, sticky="ew", padx=(10, 0))

    # Label für "Ziel-Excel-Dateipfad" erstellen
    excel_label = tk.Label(frame, text="Target Excel file path:")
    excel_label.grid(row=4, column=0, sticky="w", padx=10, pady=10)
    # Eingabefeld für den Ziel-Excel-Dateipfad erstellen
    excel_entry = tk.Entry(frame)
    if "excel_path" in config:
        excel_entry.insert(0, config["excel_path"])
    excel_entry.grid(row=4, column=1, sticky="ew", padx=10, pady=10)

    # Durchsuchen-Button erstellen
    browse_button = tk.Button(frame, text="Browse", command=browse_file)
    browse_button.grid(row=4, column=2, padx=10, pady=10)

    # Label für "Arbeitsblattname" erstellen
    worksheet_label = tk.Label(frame, text="Worksheet name:")
    worksheet_label.grid(row=5, column=0, sticky="w", padx=10, pady=10)
    # Eingabefeld für den Arbeitsblattnamen erstellen
    worksheet_entry = tk.Entry(frame)
    if "worksheet_name" in config:
        worksheet_entry.insert(0, config["worksheet_name"])
    worksheet_entry.grid(row=5, column=1, sticky="ew", padx=10, pady=10)

    # Button zum Starten der Verarbeitung erstellen
    # Die .strip()-Methode entfernt Leerzeichen am Anfang und Ende des Strings,
    # einschließlich Leerzeichen, Tabulatoren und Zeilenumbrüchen
    process_button = tk.Button(frame, text="Start processing",
                               command=lambda: [
                                    # Aktuelle Eingaben in der Konfiguration speichern
                                    save_config({
                                        "days": days_entry.get().strip(),
                                        "orders_limit": orders_entry.get().strip(),
                                        "excel_path": excel_entry.get().strip(),
                                        "worksheet_name": worksheet_entry.get().strip(),
                                    }),
                                    # Informationsanzeige leeren
                                    info_text.delete("1.0", tk.END),
                                    # Verarbeitungsstart anzeigen
                                    info_text.insert(tk.END, "Start processing... \n"),
                                    # Funktion process_orders ausführen
                                    process_orders(token_entry.get("1.0", tk.END).strip(),
                                                              days_entry.get().strip(),
                                                              orders_entry.get().strip(),
                                                              excel_entry.get().strip(),
                                                              worksheet_entry.get().strip()
                                                              )
                                ]
                               )
    process_button.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

    # Informationsanzeige erstellen
    info_frame = tk.Frame(root)
    info_frame.pack(padx=20, pady=20)

    info_label = tk.Label(info_frame, text="Information display")
    info_label.pack()

    # Einen Textbereich zur Anzeige von Informationen erstellen
    global info_text
    info_text = tk.Text(info_frame, wrap=tk.WORD, width=width, height=10, borderwidth=2)
    info_text.pack(padx=10, pady=10)

    # Kontextmenü für Rechtsklick zum Kopieren des Textes aktivieren
    info_text.bind("<Button-3>", lambda event: info_text.event_generate('<Control-a>'))
    info_text.bind("<Control-a>", lambda event: info_text.event_generate('<Copy>'))

    # Hauptereignisschleife ausführen
    root.mainloop()

# Funktion zur Verarbeitung von Bestellungen definieren
def process_orders(token, days, orders_limit, excel_path, worksheet_name):

    """Zuerst eine erste Überprüfung der eingegebenen Daten durchführen"""
    # Überprüfen, ob die Anzahl der Bestellungstage und die Anzahl der zurückzugebenden Bestellungen leer sind
    if not token or not days or not orders_limit or not excel_path or not worksheet_name:
        info_text.insert(tk.END, "All input data cannot be empty.\n")
        return

    # Überprüfen, ob die Anzahl der Bestellungstage und die Anzahl der zurückzugebenden Bestellungen Ganzzahlen sind
    if not days.isdigit() or not orders_limit.isdigit():
        info_text.insert(tk.END, "Order days and orders limit must be integers.")
        return

    # Die Anzahl der Bestellungstage und die Anzahl der zurückzugebenden Bestellungen in Ganzzahlen umwandeln
    days = int(days)
    orders_limit = int(orders_limit)

    # Versuchen, die Arbeitsmappe zu öffnen und zu überprüfen, ob die Excel-Datei und das Arbeitsblatt existieren
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook[worksheet_name]
    except FileNotFoundError:
        info_text.insert(tk.END, "The specified Excel file does not exist.\n")
        return
    except KeyError:
        info_text.insert(tk.END, "The specified worksheet does not exist.\n")
        return
    except Exception as e:
        info_text.insert(tk.END, f"Error opening Excel file: {str(e)}\n")
        return

    # Anfangsnachricht
    info_text.insert(tk.END, f"Verarbeite Bestellungen der letzten {days} Tage...\n")

    # Aktuelle Zeit als ISO 8601-formatierte Zeichenkette abrufen
    current_time = datetime.now(timezone.utc).replace(microsecond=0)  # Aktuelle Zeit
    past_24_hours = current_time - timedelta(hours=24)  # 24 Stunden zuvor
    past_x_days = current_time - timedelta(days=days)  # Anzahl der Tage zuvor

    # Zeit in das ISO 8601-Format umwandeln
    current_time_str = current_time.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    past_24_hours_str = past_24_hours.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    past_x_days_str = past_x_days.strftime("%Y-%m-%dT%H:%M:%S.000Z")

    # eBay API-Zugriffstoken festlegen
    access_token = token

    # API-Endpunkt zum Abrufen von Bestellungen definieren
    url_get_orders = "https://api.ebay.com/sell/fulfillment/v1/order"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }

    # Liste zum Speichern von Bestellinformationen definieren
    orders_list = []

    params_get_orders = {
        "filter": f"creationdate:[{past_x_days_str}..{current_time_str}]",
        "limit": orders_limit  # Anzahl der zurückzugebenden Bestellungen anpassen
    }

    # API-Anfrage zum Abrufen von Bestellungen senden
    response_get_orders = requests.get(url_get_orders, headers=headers, params=params_get_orders)

    # Alle Bestellungen mit getOrders abrufen
    if response_get_orders.status_code == 200:
        orders = response_get_orders.json().get("orders", [])
        # print(orders) # Zum Debuggen: Bestellinformationen anzeigen

        # Jede Bestellung mit getOrder durchlaufen
        for order in orders:
            order_id = order.get("orderId")
            # Empfänger- und Adressinformationen der Bestellung abrufen
            url_get_order = f"https://api.ebay.com/sell/fulfillment/v1/order/{order_id}"
            response_get_order = requests.get(url_get_order, headers=headers)

            if response_get_order.status_code == 200:
                order_details = response_get_order.json()
                print(order_details)
                shipping_step = order_details.get('fulfillmentStartInstructions', [{}])[0].get('shippingStep', {})
                ship_to = shipping_step.get('shipTo', {})
                # creationDate = order_details.get('creationDate', 'Nicht angegeben')[:10]  # Nur das Datumsteil abrufen, also bis zum Tag abschneiden
                # Das vollständige Datum einschließlich der Uhrzeit abrufen, um Bestellungen am selben Tag von der ältesten zur neuesten zu sortieren
                creationDate = order_details.get('creationDate', 'Nicht angegeben')
                order_fulfillment_status = order_details.get('orderFulfillmentStatus', 'Nicht angegeben')
                cancel_status = order_details.get('cancelStatus', {}).get('cancelState', 'Nicht angegeben')
                full_name = ship_to.get('fullName', 'Nicht angegeben')
                contact_address = ship_to.get('contactAddress', {})
                Strasse1 = contact_address.get('addressLine1', 'Nicht angegeben')
                Strasse2 = contact_address.get('addressLine2', '')
                # Wenn Strasse2 leer ist, dann ist Strasse gleich Strasse1, ansonsten ist Strasse gleich Strasse1 + ' (' + Strasse2 + ')'
                Strasse = Strasse1 if not Strasse2 else Strasse1 + ' (' + Strasse2 + ')'
                city = contact_address.get('city', 'Nicht angegeben')
                PLZ = contact_address.get('postalCode', 'Nicht angegeben')
                phone_number_dict = ship_to.get('primaryPhone', {})
                phone_number = phone_number_dict.get('phoneNumber', 'Nicht angegeben')
                email = ship_to.get('email', 'Nicht angegeben')
                buyer_username = order_details.get('buyer', {}).get('username', 'Nicht angegeben')  # Benutzername des Käufers abrufen

                # Eine Bestellung kann mehrere Artikel enthalten, daher müssen die Artikelinformationen durchlaufen werden
                item_info = order_details.get('lineItems', [{}])
                for item in item_info:
                    sku = item.get('sku', 'Nicht angegeben')
                    quantity = item.get('quantity', 'Nicht angegeben')
                    # Der von EBAY erhaltene Preis ist ein String und muss in einen Float-Typ umgewandelt werden, um mathematische Berechnungen durchführen zu können
                    # price_str = item.get('lineItemCost', {}).get('value', 'Nicht angegeben')  # Ursprünglicher Preis
                    # Die folgende Zeile gibt den rabattierten Preis zurück, also den endgültigen Verkaufspreis
                    price_str = item.get('total', {}).get('value', '未提供')
                    # Preis in den Float-Typ umwandeln
                    price = float(price_str)

                    # Bestellinformationen als Wörterbuch speichern
                    order_info = {
                        "order_id": order_id,
                        "creationDate": creationDate,
                        "order_fulfillment_status": order_fulfillment_status,
                        "cancel_status": cancel_status,
                        "full_name": full_name,
                        "Strasse": Strasse,
                        "city": city,
                        "PLZ": PLZ,
                        "phone_number": phone_number,
                        "email": email,
                        "buyer_username": buyer_username,  # Benutzername des Käufers hinzufügen
                        "sku": sku,
                        "quantity": quantity,
                        "price": price
                    }

                    # Bestellinformationen-Wörterbuch zur Liste hinzufügen
                    orders_list.append(order_info)

            else:
                print(f"Fehler beim Abrufen der Bestellung {order_id}, Statuscode: {response_get_order.status_code}")
                info_text.insert(tk.END, f"Fehler beim Abrufen der Bestellung {order_id}, Statuscode: {response_get_order.status_code}\n")
    else:
        print(f"Fehler beim Abrufen der Bestellliste, Statuscode: {response_get_orders.status_code}")
        info_text.insert(tk.END, f"Fehler beim Abrufen der Bestellliste, Statuscode: {response_get_orders.status_code}\n")

    # Bestellinformationen in der Konsole ausgeben
    print("Alle abgerufenen Bestellungen:")
    for order_info in orders_list:
        print(order_info)
    print()  # Leerzeile zur besseren Lesbarkeit

    # Bestellinformationen im UI-Informationsfenster anzeigen
    info_text.insert(tk.END, "Alle abgerufenen Bestellungen:\n")
    for order_info in orders_list:
        info_text.insert(tk.END, str(order_info) + "\n")
    info_text.insert(tk.END, "\n")  # Leerzeile zur besseren Lesbarkeit

    """
    2. Die erhaltene Informationsliste weiterverarbeiten, zuerst stornierte Bestellungen entfernen, dann nach Datum von der ältesten zur neuesten sortieren
    """

    # Eine neue Liste definieren, um nicht stornierte Bestellinformationen zu speichern
    uncanceled_orders_list = []

    # 迭代订单信息列表
    for order_info in orders_list:
        # Überprüfen, ob der Stornierungsstatus der Bestellung "CANCELED" ist
        if order_info["cancel_status"] != "CANCELED":
            # Wenn nicht "CANCELED", die Bestellinformationen zur neuen Liste hinzufügen
            uncanceled_orders_list.append(order_info)

    # Gefilterte Liste der Bestellinformationen    # Gefilterte Bestellungen ausgeben
    print("Bereinigte Bestellliste (stornierte Bestellungen wurden entfernt):")
    for order_info in uncanceled_orders_list:
        print(order_info)
    print()  # Leerzeile zur besseren Lesbarkeit

    # Eine Funktion definieren, um das Erstellungsdatum einer Bestellung für die Sortierung zu erhalten
    def get_creation_date(order_info):
        return order_info["creationDate"]

    # Die Liste der nicht stornierten Bestellungen nach Erstellungsdatum von der ältesten zur neuesten sortieren
    sorted_uncanceled_orders_list = sorted(uncanceled_orders_list, key=get_creation_date)

    # Durch die sortierte Bestellliste iterieren, nur das Datum ohne Uhrzeit behalten, da es sonst zu Formatfehlern beim Schreiben in die Excel-Tabelle kommt
    for order_info in sorted_uncanceled_orders_list:
        order_info["creationDate"] = order_info["creationDate"][:10]  # Nur das Datum behalten

    # Sortierte Liste der Bestellinformationen    # Sortierte Bestellungen ausgeben
    print("Nach Datum sortierte Bestellungen (älteste zuerst):")
    for order_info in sorted_uncanceled_orders_list:
        print(order_info)
    print()  # Leerzeile zur besseren Lesbarkeit

    """
    Drei. jetzt müssen wir weiterverarbeiten, um zu prüfen, ob es sich um eine Dusche HLMR, DR, DBL oder CL handelt
    1. Teile die SKU in drei Teile auf: Buchstaben, erste zwei Ziffern, letzte zwei Ziffern
    2. Überprüfe, ob der Buchstabe in der Liste enthalten ist, die ersten beiden Ziffern und die letzten beiden Ziffern sind identisch
{{ ... }}
    4. Wenn der Buchstabe HLMR, DR, DBL oder CL ist, aber die ersten beiden Ziffern und die letzten beiden Ziffern nicht identisch sind, dann teile die Zeile in zwei Zeilen auf, eine Zeile mit der SKU Buchstabe + erste zwei Ziffern, die andere Zeile mit der SKU Buchstabe + letzte zwei Ziffern, Preis ändere in 0
    """

    # Neue Liste für die verarbeiteten Bestellungen erstellen
    processed_orders_list = []

    # 迭代订单信息列表
    for order_info in sorted_uncanceled_orders_list:
        sku = order_info["sku"]
        # print(sku)
        if sku:

            # Zuerst prüfen, ob die SKU mit 'NF' endet
            if sku.endswith('NF'):
                # Erste Zeile: Ursprüngliche SKU behalten (ohne 'NF')
                first_order = order_info.copy()
                first_order["sku"] = sku[:-2]  # 'NF' entfernen
                processed_orders_list.append(first_order)
                
                # Zweite Zeile: WA30H verwenden
                second_order = order_info.copy()
                second_order["sku"] = "DWR30"
                second_order["price"] = 0  # Preis auf 0 setzen
                processed_orders_list.append(second_order)

            # Wenn nicht mit 'NF' endend, die ursprüngliche Logik ausführen
            # Prüfen, ob die SKU die Bedingung erfüllt: Nach dem viertletzten Zeichen nur Buchstaben, dritt- und viertletzte Ziffern, letzte zwei Ziffern
            elif sku[:-4].isalpha() and sku[-4:-2].isdigit() and sku[-2:].isdigit():
                # SKU in drei Teile aufteilen
                letters = sku[:-4]
                first_two_digits = sku[
                                   -4:-2]  # Extrahiert die Zeichen an Position -4 und -3 (exklusive -2)
                last_two_digits = sku[-2:]

                # Wenn first_two_digits 10, 11 oder 12 ist, eine Null anhängen
                if first_two_digits in ['10', '11', '12']:
                    first_two_digits = first_two_digits + '0'

                # Wenn last_two_digits 10, 11 oder 12 ist, eine Null anhängen
                if last_two_digits in ['10', '11', '12']:
                    last_two_digits = last_two_digits + '0'

                # Prüfen, ob die Buchstaben in der angegebenen Liste sind und ob die ersten beiden Ziffern mit den letzten beiden übereinstimmen
                if letters in ["HLMR", "DR", "CL", "DBL"]:

                    if first_two_digits == last_two_digits:
                        # Bestellinformationen anpassen
                        order_info["sku"] = letters + first_two_digits
                        order_info["quantity"] = 2 * order_info["quantity"]  # Menge verdoppeln
                        processed_orders_list.append(order_info.copy())

                    else:
                        # In zwei Zeilen verarbeiten
                        order_info["sku"] = letters + first_two_digits
                        processed_orders_list.append(order_info.copy())  # Erste Zeile hinzufügen
                        order_info["sku"] = letters + last_two_digits
                        order_info["price"] = 0  # Preis auf 0 setzen
                        processed_orders_list.append(order_info.copy())  # Zweite Zeile hinzufügen

                else:
                    # Wenn der Typ nicht einer der 4 Typen ist, unverändert lassen
                    processed_orders_list.append(order_info)

            else:
                # Wenn die SKU die Bedingungen nicht erfüllt, Bestellinformationen unverändert lassen
                processed_orders_list.append(order_info)
        else:
            # Wenn keine SKU vorhanden ist, Bestellinformationen unverändert lassen
            processed_orders_list.append(order_info)

    # Verarbeitete Bestellinformationen ausgeben
    print("Die folgenden Bestellungen wurden für den Export nach Excel vorbereitet:")
    for order_info in processed_orders_list:
        print(order_info)
    print()  # Leerzeile zur besseren Lesbarkeit

    """
    Vier。 Informationen in EXCEL-Datei schreiben
    1. Excel-Tabelle durchsuchen und Bestellnummern (Spalte H) in einer Menge speichern
    2. Bestellungen aus der Menge processed_orders_list entfernen, die bereits in der Excel-Tabelle vorhanden sind
    3. Restliche Bestellungen in Excel-Tabelle schreiben
    """

    # Funktion zum Laden der Excel-Tabelle und Rückgabe der Bestellnummern
    def load_order_ids_from_excel(file_path, sheet_name):  # Parameter: Excel-Dateipfad und Arbeitsblattname
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        order_ids = set()
        for row in sheet.iter_rows(min_row=2, max_col=8, max_row=sheet.max_row, values_only=True):
            order_ids.add(row[7])  # Bestellnummer ist in Spalte H (Index 7)
        return order_ids

    # Bestellnummern aus der Excel-Tabelle laden
    excel_order_ids = load_order_ids_from_excel(excel_path, worksheet_name)

    # Bereits in der Excel-Tabelle vorhandene Bestellungen aus processed_orders_list entfernen
    processed_orders_list = [order for order in processed_orders_list if order['order_id'] not in excel_order_ids]

    # Bereinigte Bestellliste ausgeben
    print("Bereinigte Bestellliste (bereits in Excel vorhandene Bestellungen wurden entfernt):")
    for order_info in processed_orders_list:
        print(order_info)

    # Funktion zum Schreiben in die Excel-Datei definieren
    def write_orders_to_excel(orders_list, file_path):  # Eingabeparameter: Bestellliste und Excel-Dateipfad
        # Excel-Datei öffnen
        workbook = openpyxl.load_workbook(file_path)

        # Arbeitsblatt mit dem angegebenen Namen auswählen
        sheet = workbook[worksheet_name]  # Direkter Zugriff auf das Arbeitsblatt, um DeprecationWarning zu vermeiden

        # Nächste leere Zeile in Spalte A finden, um mit dem Schreiben zu beginnen
        next_row = 1
        while sheet.cell(row=next_row, column=1).value:
            next_row += 1

        # Bestellliste durchlaufen und Daten schreiben
        for order in orders_list:
            # Bestellinformationen in Excel schreiben
            sheet.cell(row=next_row, column=1).value = order['creationDate']  # Bestelldatum
            sheet.cell(row=next_row, column=2).value = 'Ebay'  # Plattform
            sheet.cell(row=next_row, column=3).value = order['quantity']  # Menge
            sheet.cell(row=next_row, column=4).value = order['price']  # Preis
            sheet.cell(row=next_row, column=5).value = order['sku']  # Artikelnummer
            sheet.cell(row=next_row, column=6).value = 'Wuppertal'  # Standort
            sheet.cell(row=next_row, column=7).value = 'Ebay'  # Verkaufsplattform
            sheet.cell(row=next_row, column=8).value = order['order_id']  # Bestellnummer
            sheet.cell(row=next_row, column=10).value = order['buyer_username']  # Käufername
            sheet.cell(row=next_row, column=11).value = order['email']  # E-Mail-Adresse
            sheet.cell(row=next_row, column=12).value = order['phone_number']  # Telefonnummer
            sheet.cell(row=next_row, column=13).value = 'Versand'  # Versand
            sheet.cell(row=next_row, column=14).value = order['full_name']  # Empfängername
            sheet.cell(row=next_row, column=15).value = order['Strasse']  # Straße
            sheet.cell(row=next_row, column=16).value = order['PLZ']  # Postleitzahl
            sheet.cell(row=next_row, column=17).value = order['city']  # Stadt
            # Neue Zeile einfügen
            sheet.insert_rows(next_row + 1)
            next_row += 1  # Zur nächsten Zeile wechseln

            # Zeilennummer der soeben geschriebenen Datumszelle abrufen
            date_row = next_row - 1

            # Textwert des soeben geschriebenen Datums in ein Datumsobjekt umwandeln und in das Format TT.MM.JJJJ ändern
            date_column = 1  # Spalte des Datums
            date_cell = sheet.cell(row=date_row, column=date_column)
            date_cell.value = datetime.strptime(str(date_cell.value), "%Y-%m-%d")
            date_cell.number_format = "DD.MM.YY"

        # Excel-Datei speichern
        workbook.save(file_path)
        info_text.insert(tk.END, f"\nBestellverarbeitung erfolgreich abgeschlossen!")

    # Funktion aufrufen, um Bestellungen in die Excel-Datei zu schreiben
    write_orders_to_excel(processed_orders_list, excel_path)


if __name__ == "__main__":
    main()





